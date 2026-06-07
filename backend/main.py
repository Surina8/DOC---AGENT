from fastapi import FastAPI, UploadFile, File, Form, Depends, Body
from fastapi.middleware.cors import CORSMiddleware
from sqlalchemy.orm import Session
from dotenv import load_dotenv
from datetime import datetime
import tempfile, os, json, fitz

from pdf_reader import extract_text
from prompt_builder import build_suggest_prompt, build_extract_prompt
from llm_client import call_llm
from confidence import find_coordinates_and_confidence
from database import get_db, SessionLocal
from models import Document, Extraction, Field, User
from auth import (
    hash_password, verify_password, create_access_token,
    get_current_user
)

from pydantic import BaseModel
from typing import List, Optional
from models import FieldTemplate, TemplateField, Batch

from embeddings import generate_embedding

import threading
from sqlalchemy.orm import joinedload
from models import Batch

from fastapi.responses import StreamingResponse, FileResponse
import io

import math

load_dotenv()

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:3000"],
    allow_methods=["*"],
    allow_headers=["*"],
)

PDF_DIR = "pdfs"
os.makedirs(PDF_DIR, exist_ok=True)


@app.get("/")
def root():
    return {"status": "DocAgent backend teče"}


# Avtentikacija

class RegisterRequest(BaseModel):
    email: str
    password: str
    full_name: Optional[str] = None


class LoginRequest(BaseModel):
    email: str
    password: str


@app.post("/api/auth/register")
def register(payload: RegisterRequest, db: Session = Depends(get_db)):
    if len(payload.password) < 6:
        return {"error": "Geslo mora imeti vsaj 6 znakov"}

    existing = db.query(User).filter(User.email == payload.email.lower()).first()
    if existing:
        return {"error": "Uporabnik s tem emailom že obstaja"}

    user = User(
        email=payload.email.lower(),
        hashed_password=hash_password(payload.password),
        full_name=payload.full_name,
    )
    db.add(user)
    db.commit()
    db.refresh(user)

    token = create_access_token(str(user.id), user.email)
    return {
        "access_token": token,
        "token_type": "bearer",
        "user": {
            "id": str(user.id),
            "email": user.email,
            "full_name": user.full_name,
            "role": user.role,
        }
    }


@app.post("/api/auth/login")
def login(payload: LoginRequest, db: Session = Depends(get_db)):
    user = db.query(User).filter(User.email == payload.email.lower()).first()
    if not user or not verify_password(payload.password, user.hashed_password):
        return {"error": "Napačen email ali geslo"}

    if not user.is_active:
        return {"error": "Račun ni aktiven"}

    user.last_login = datetime.utcnow()
    db.commit()

    token = create_access_token(str(user.id), user.email)
    return {
        "access_token": token,
        "token_type": "bearer",
        "user": {
            "id": str(user.id),
            "email": user.email,
            "full_name": user.full_name,
            "role": user.role,
        }
    }


@app.get("/api/auth/me")
def get_me(user: User = Depends(get_current_user)):
    return {
        "id": str(user.id),
        "email": user.email,
        "full_name": user.full_name,
        "role": user.role,
        "created_at": user.created_at.isoformat(),
    }


@app.get("/api/pdfs/{document_id}")
def serve_pdf(
    document_id: str,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """PDF samo za lastnika dokumenta."""
    from fastapi import HTTPException
    doc = db.query(Document).filter(
        Document.id == document_id,
        Document.user_id == user.id,
    ).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Dokument ne obstaja")
    if not os.path.exists(doc.pdf_path):
        raise HTTPException(status_code=404, detail="PDF datoteka ne obstaja")
    return FileResponse(doc.pdf_path, media_type="application/pdf")


@app.post("/api/suggest-fields")
async def suggest_fields(
    file: UploadFile = File(...),
    user: User = Depends(get_current_user),
):
    with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
        tmp.write(await file.read())
        tmp_path = tmp.name
    try:
        text = extract_text(tmp_path)
        prompt = build_suggest_prompt(text)
        raw = call_llm(prompt)

        print("=== SUGGEST RAW RESPONSE ===")
        print(repr(raw))  # repr pokaže tudi prazne stringe, novike, itd.
        print("=" * 30)

        result = json.loads(raw)
        return result
    finally:
        os.remove(tmp_path)


@app.post("/api/extract")
async def extract(
    file: UploadFile = File(...),
    fields: str = Form("[]"),
    template_id: str = Form(""),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    pdf_filename = f"{file.filename}"
    pdf_path = os.path.join(PDF_DIR, pdf_filename)

    content = await file.read()
    with open(pdf_path, "wb") as f:
        f.write(content)

    try:
        # 1) Pridobi tekst in št. strani
        text = extract_text(pdf_path)
        embedding = generate_embedding(text)
        print(f"USTVARJEN embedding: {len(embedding)} dimenzij")

        with fitz.open(pdf_path) as pdf_doc:
            total_pages = pdf_doc.page_count

        fields_list = json.loads(fields)
        print("FIELDS PREJETE:", len(fields_list), "polj")

        # 2) Ustvari Document zapis v bazi
        document = Document(
            user_id=user.id,
            filename=file.filename,
            pdf_path=pdf_path,
            total_pages=total_pages,
            document_text=text,
            embedding=embedding,
            template_id=template_id if template_id else None,
            status="pending"
        )
        db.add(document)
        db.flush()  # Dobimo document.id še preden commit
        print(f"USTVARJEN Document: {document.id}")

        # 3) Pokliči LLM
        prompt = build_extract_prompt(fields_list, text)
        raw = call_llm(prompt)

        print("==== RAW LLM ODGOVOR ====")
        print(raw)
        print("=========================")

        extraction_data = json.loads(raw)

        # 4) Najdi koordinate in confidence
        results = find_coordinates_and_confidence(pdf_path, extraction_data)
        # DEBUG: pokaži OCR bloke in iskanja za prvih nekaj polj
        from confidence import debug_print_blocks
        debug_print_blocks(pdf_path)
        for key, data in list(extraction_data.items())[:3]:
            if isinstance(data, dict) and data.get("value"):
                print(f"\n>> Iščem: '{data['value']}'")
                debug_print_blocks(pdf_path, data['value'])
        # 5) Izračunaj povprečno confidence
        confidences = [
            v["confidence"] for v in results.values()
            if v.get("confidence") is not None
        ]
        avg_conf = sum(confidences) / len(confidences) if confidences else 0.0

        # 6) Ustvari Extraction zapis
        extraction = Extraction(
            document_id=document.id,
            model_used="openai/gpt-4o-mini",
            avg_confidence=avg_conf,
            raw_llm_response=extraction_data
        )
        db.add(extraction)
        db.flush()
        print(f"USTVARJEN Extraction: {extraction.id} (avg_conf={avg_conf:.2f})")

        # 7) Ustvari Field zapise za vsako polje
        for field_key, field_data in results.items():
            field_record = Field(
                extraction_id=extraction.id,
                field_key=field_key,
                field_value=field_data.get("value"),
                source_text=field_data.get("source_text"),
                confidence=field_data.get("confidence") or 0.0,
                page_number=field_data.get("page"),
                rectangles=field_data.get("rectangles", [])
            )
            db.add(field_record)

        db.commit()
        print(f"SHRANJENO V BAZO: {len(results)} polj")

        return {
            "document_id": str(document.id),
            "extraction_id": str(extraction.id),
            "pdf_url": f"http://localhost:8000/api/pdfs/{document.id}",
            "results": results
        }

    except Exception as e:
        db.rollback()
        print(f"NAPAKA v extract: {e}")
        import traceback
        traceback.print_exc()
        return {"error": str(e)}
    
@app.get("/api/documents")
def list_documents(
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """
    Vrne seznam vseh dokumentov, sortiran po datumu naloženja (najnovejši prvi).
    Vsak dokument vključuje povzetek zadnje ekstrakcije.
    """
    docs = db.query(Document).filter(
        Document.user_id == user.id
    ).order_by(Document.upload_date.desc()).all()

    result = []
    for doc in docs:
        # Vzemi zadnjo ekstrakcijo (po datumu)
        latest_extraction = None
        if doc.extractions:
            latest_extraction = max(doc.extractions, key=lambda e: e.extraction_date)

        result.append({
            "id": str(doc.id),
            "filename": doc.filename,
            "upload_date": doc.upload_date.isoformat(),
            "total_pages": doc.total_pages,
            "status": doc.status,
            "field_count": len(latest_extraction.fields) if latest_extraction else 0,
            "avg_confidence": latest_extraction.avg_confidence if latest_extraction else None,
        })

    return {"documents": result}


@app.get("/api/documents/{document_id}")
def get_document(
    document_id: str,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """
    Vrne podrobnosti enega dokumenta z vsemi polji zadnje ekstrakcije.
    Format je enak kot `/api/extract` response, da lahko frontend
    Review stran uporabi isti UI.
    """
    doc = db.query(Document).filter(
        Document.id == document_id,
        Document.user_id == user.id,
    ).first()
    if not doc:
        return {"error": "Dokument ne obstaja"}

    if not doc.extractions:
        return {"error": "Dokument še nima ekstrakcij"}

    latest = max(doc.extractions, key=lambda e: e.extraction_date)

    # Sestavi results dict v isti obliki kot extract endpoint
    results = {}
    for field in latest.fields:
        results[field.field_key] = {
            "value": field.field_value,
            "source_text": field.source_text,
            "confidence": field.confidence,
            "page": field.page_number,
            "rectangles": field.rectangles or [],
        }

    return {
        "document_id": str(doc.id),
        "extraction_id": str(latest.id),
        "filename": doc.filename,
        "upload_date": doc.upload_date.isoformat(),
        "total_pages": doc.total_pages,
        "status": doc.status,
        "confirmed_at": latest.confirmed_at.isoformat() if latest.confirmed_at else None,
        "corrections_count": latest.corrections_count,
        "pdf_url": f"http://localhost:8000/api/pdfs/{doc.id}",
        "results": results,
    }


@app.delete("/api/documents/{document_id}")
def delete_document(
    document_id: str,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """
    Izbriše dokument in vse povezane podatke (extraction, fields).
    Cascade delete v models.py poskrbi za otroke.
    """
    doc = db.query(Document).filter(
        Document.id == document_id,
        Document.user_id == user.id,
    ).first()
    if not doc:
        return {"error": "Dokument ne obstaja"}

    pdf_path = doc.pdf_path

    db.delete(doc)
    db.commit()

    # Izbriši še PDF datoteko z diska
    if os.path.exists(pdf_path):
        try:
            os.remove(pdf_path)
        except Exception as e:
            print(f"Opozorilo: PDF ni mogel biti izbrisan: {e}")

    return {"status": "izbrisano"}


@app.get("/api/archive")
def list_archive(
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """
    Vrne vse POTRJENE dokumente (z extraction.confirmed_at != None).
    Razlikuje se od /api/documents, ki vrne vse (vključno z nepotrjenimi).
    """
    docs = db.query(Document).filter(
        Document.user_id == user.id
    ).order_by(Document.upload_date.desc()).all()

    result = []
    for doc in docs:
        if not doc.extractions:
            continue
        latest = max(doc.extractions, key=lambda e: e.extraction_date)
        if not latest.confirmed_at:
            continue  # samo potrjeni

        result.append({
            "id": str(doc.id),
            "extraction_id": str(latest.id),
            "filename": doc.filename,
            "upload_date": doc.upload_date.isoformat(),
            "confirmed_at": latest.confirmed_at.isoformat(),
            "total_pages": doc.total_pages,
            "field_count": len(latest.fields),
            "avg_confidence": latest.avg_confidence,
            "corrections_count": latest.corrections_count or 0,
            "template_id": str(doc.template_id) if doc.template_id else None,
        })

    return {"documents": result}


def _safe_filename(name: str, ext: str) -> str:
    """
    Naredi ASCII-safe ime datoteke za HTTP Content-Disposition.
    Slovenske črke → ASCII ekvivalenti, posebni znaki → _.
    """
    import os, re, unicodedata
    base = os.path.splitext(name)[0]  # strip .pdf
    # Slovenske črke → ASCII
    base = unicodedata.normalize('NFKD', base).encode('ascii', 'ignore').decode('ascii')
    # Vsi non-alphanumeric, razen _ - → _
    base = re.sub(r'[^\w\-]', '_', base)
    # Več _ skupaj → en _
    base = re.sub(r'_+', '_', base).strip('_')
    if not base:
        base = "dokument"
    return f"{base}.{ext}"


@app.get("/api/documents/{document_id}/export")
def export_single_document(
    document_id: str,
    format: str = "json",
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """
    Izvozi en dokument v izbranem formatu: json | csv | excel | txt
    Vrne SAMO raw podatke (field_key → value), brez confidence/metadat.
    """
    doc = db.query(Document).filter(
        Document.id == document_id,
        Document.user_id == user.id,
    ).first()
    if not doc or not doc.extractions:
        return {"error": "Dokument ne obstaja ali nima ekstrakcij"}

    latest = max(doc.extractions, key=lambda e: e.extraction_date)

    # Samo key → value mapping
    fields_data = {f.field_key: (f.field_value or "") for f in latest.fields}

    if format == "json":
        content = json.dumps(fields_data, ensure_ascii=False, indent=2)
        media_type = "application/json"
        ext = "json"

    elif format == "csv":
        import csv as csv_module
        buf = io.StringIO()
        writer = csv_module.writer(buf)
        writer.writerow(["field_key", "value"])
        for k, v in fields_data.items():
            writer.writerow([k, v])
        content = buf.getvalue()
        media_type = "text/csv"
        ext = "csv"

    elif format == "excel":
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = "Podatki"

        ws.append(["Polje", "Vrednost"])
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="3B82F6")

        for k, v in fields_data.items():
            ws.append([k, v])

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{_safe_filename(doc.filename, "xlsx")}"'}
        )

    elif format == "txt":
        lines = [f"{k}: {v}" for k, v in fields_data.items()]
        content = "\n".join(lines)
        media_type = "text/plain; charset=utf-8"
        ext = "txt"

    else:
        return {"error": f"Format '{format}' ni podprt. Uporabi: json, csv, excel, txt"}

    return StreamingResponse(
        io.BytesIO(content.encode("utf-8")),
        media_type=media_type,
        headers={"Content-Disposition": f'attachment; filename="{_safe_filename(doc.filename, ext)}"'}
    )


class BulkExportRequest(BaseModel):
    document_ids: List[str]
    format: str = "json"


@app.post("/api/archive/export")
def bulk_export(
    payload: BulkExportRequest,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """
    Bulk export izbranih dokumentov v izbranem formatu.
    """
    if not payload.document_ids:
        return {"error": "Ni izbranih dokumentov"}

    docs_data = []
    for doc_id in payload.document_ids:
        doc = db.query(Document).filter(
            Document.id == doc_id,
            Document.user_id == user.id,
        ).first()
        if not doc or not doc.extractions:
            continue
        latest = max(doc.extractions, key=lambda e: e.extraction_date)

        # Samo key → value (brez confidence, user_corrected itd.)
        fields = {f.field_key: (f.field_value or "") for f in latest.fields}

        docs_data.append({
            "filename": doc.filename,
            "fields": fields,
        })

    if payload.format == "json":
        content = json.dumps({"documents": docs_data}, ensure_ascii=False, indent=2)
        return StreamingResponse(
            io.BytesIO(content.encode("utf-8")),
            media_type="application/json",
            headers={"Content-Disposition": 'attachment; filename="arhiv_export.json"'}
        )

    elif payload.format == "csv":
        import csv as csv_module
        # Zberi vse unique field keys
        all_keys = set()
        for d in docs_data:
            all_keys.update(d["fields"].keys())
        sorted_keys = sorted(all_keys)

        buf = io.StringIO()
        writer = csv_module.writer(buf)
        writer.writerow(["filename"] + sorted_keys)
        for d in docs_data:
            row = [d["filename"]]
            for k in sorted_keys:
                row.append(d["fields"].get(k, ""))
            writer.writerow(row)
        return StreamingResponse(
            io.BytesIO(buf.getvalue().encode("utf-8")),
            media_type="text/csv",
            headers={"Content-Disposition": 'attachment; filename="arhiv_export.csv"'}
        )

    elif payload.format == "excel":
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = "Pregled"

        all_keys = set()
        for d in docs_data:
            all_keys.update(d["fields"].keys())
        sorted_keys = sorted(all_keys)

        headers = ["Datoteka"] + sorted_keys
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="3B82F6")

        for d in docs_data:
            row = [d["filename"]]
            for k in sorted_keys:
                row.append(d["fields"].get(k, ""))
            ws.append(row)

        # Auto-width
        for col in ws.columns:
            max_len = max(len(str(cell.value)) for cell in col if cell.value)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": 'attachment; filename="arhiv_export.xlsx"'}
        )

    elif payload.format == "txt":
        lines = []
        for d in docs_data:
            lines.append(f"=== {d['filename']} ===")
            for k in sorted(d["fields"].keys()):
                lines.append(f"{k}: {d['fields'][k]}")
            lines.append("")
        content = "\n".join(lines)
        return StreamingResponse(
            io.BytesIO(content.encode("utf-8")),
            media_type="text/plain; charset=utf-8",
            headers={"Content-Disposition": 'attachment; filename="arhiv_export.txt"'}
        )

    else:
        return {"error": f"Format '{payload.format}' ni podprt"}


@app.post("/api/extractions/{extraction_id}/confirm")
def confirm_extraction(
    extraction_id: str,
    corrections: dict = Body(...),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """
    Označi extraction kot 'human-reviewed' in shrani popravke.
    """
    extraction = db.query(Extraction).filter(Extraction.id == extraction_id).first()
    if not extraction:
        return {"error": "Extraction ne obstaja"}

    # Preverba lastništva preko dokumenta
    if extraction.document.user_id != user.id:
        return {"error": "Nimaš dostopa do tega dokumenta"}

    changes = []
    for field in extraction.fields:
        if field.field_key not in corrections:
            continue
        new_val = corrections[field.field_key]
        # Normaliziraj None / prazen string
        if new_val is None:
            new_val = ""
        new_val = str(new_val)
        current_val = field.field_value or ""

        if new_val != current_val:
            # Arhiviraj original (samo prvič — če je že popravljen, ne prepiši)
            if not field.user_corrected:
                field.original_value = field.field_value
            field.field_value = new_val
            field.user_corrected = True
            changes.append({
                "field_key": field.field_key,
                "original": current_val,
                "corrected": new_val,
            })

    # Označi extraction kot potrjen
    extraction.confirmed_at = datetime.utcnow()
    extraction.corrections_count = (extraction.corrections_count or 0) + len(changes)

    # Posodobi tudi document status
    if extraction.document:
        extraction.document.status = "reviewed"

    db.commit()
    print(f"POTRJEN Extraction: {extraction_id[:8]} — {len(changes)} korekcij")

    return {
        "status": "confirmed",
        "extraction_id": str(extraction.id),
        "confirmed_at": extraction.confirmed_at.isoformat(),
        "corrections": changes,
        "total_corrections": extraction.corrections_count,
    }


class TemplateFieldInput(BaseModel):
    field_key: str
    field_description: str


class TemplateCreateRequest(BaseModel):
    name: str
    description: Optional[str] = None
    document_type: Optional[str] = None
    fields: List[TemplateFieldInput]


@app.post("/api/templates")
def create_template(
    payload: TemplateCreateRequest,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Ustvari nov field template iz uporabnikovih trenutnih polj."""
    template = FieldTemplate(
        user_id=user.id,
        name=payload.name,
        description=payload.description,
        document_type=payload.document_type,
    )
    db.add(template)
    db.flush()

    for idx, field in enumerate(payload.fields):
        tf = TemplateField(
            template_id=template.id,
            field_key=field.field_key,
            field_description=field.field_description,
            order_index=idx,
        )
        db.add(tf)

    db.commit()
    print(f"USTVARJEN Template: {template.id} ({template.name}) z {len(payload.fields)} polji")

    return {
        "id": str(template.id),
        "name": template.name,
        "field_count": len(payload.fields),
    }


@app.get("/api/templates")
def list_templates(
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Vrne seznam vseh template-ov."""
    templates = db.query(FieldTemplate).filter(
        FieldTemplate.user_id == user.id
    ).order_by(
        FieldTemplate.usage_count.desc(),
        FieldTemplate.created_date.desc()
    ).all()

    result = []
    for t in templates:
        result.append({
            "id": str(t.id),
            "name": t.name,
            "description": t.description,
            "document_type": t.document_type,
            "created_date": t.created_date.isoformat(),
            "usage_count": t.usage_count,
            "field_count": len(t.template_fields),
        })

    return {"templates": result}


@app.get("/api/templates/{template_id}")
def get_template(
    template_id: str,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Vrne en template z vsemi polji."""
    template = db.query(FieldTemplate).filter(
        FieldTemplate.id == template_id,
        FieldTemplate.user_id == user.id,
    ).first()

    if not template:
        return {"error": "Template ne obstaja"}

    # Povečaj usage_count - vsako branje je "uporaba"
    # (kasneje lahko premakneš v ločen endpoint /use)
    template.usage_count += 1
    db.commit()

    return {
        "id": str(template.id),
        "name": template.name,
        "description": template.description,
        "document_type": template.document_type,
        "created_date": template.created_date.isoformat(),
        "usage_count": template.usage_count,
        "fields": [
            {
                "key": tf.field_key,
                "description": tf.field_description,
            }
            for tf in template.template_fields
        ],
    }


@app.delete("/api/templates/{template_id}")
def delete_template(
    template_id: str,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Izbriše template. Cascade izbriše tudi vsa template_fields."""
    template = db.query(FieldTemplate).filter(
        FieldTemplate.id == template_id,
        FieldTemplate.user_id == user.id,
    ).first()

    if not template:
        return {"error": "Template ne obstaja"}

    db.delete(template)
    db.commit()

    return {"status": "izbrisano"}

@app.post("/api/find-similar")
async def find_similar(
    file: UploadFile = File(...),
    limit: int = Form(5),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """
    Najde top K najbolj podobnih dokumentov za podan PDF.
    Uporablja kosinusno podobnost na embedding-ih (pgvector).

    Vrne tudi template informacije, da lahko frontend predlaga
    isti template kot je bil uporabljen pri podobnih dokumentih.
    """
    # Začasno shrani PDF, ekstrahiraj tekst, generiraj embedding
    with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
        tmp.write(await file.read())
        tmp_path = tmp.name

    try:
        text = extract_text(tmp_path)
        if not text or not text.strip():
            return {"similar_documents": [], "warning": "Dokument nima tekstovne vsebine"}

        query_embedding = generate_embedding(text)
        print(f"FIND-SIMILAR: generiran embedding za {file.filename}")

        # pgvector kosinusna podobnost preko ORM
        # cosine_distance: 0 = identično, 2 = nasprotno
        # similarity = 1 - distance: 1 = identično, 0 = popolnoma drugačno
        results = (
            db.query(
                Document,
                (1 - Document.embedding.cosine_distance(query_embedding)).label("similarity")
            )
            .filter(Document.embedding.is_not(None))
            .filter(Document.user_id == user.id)
            .order_by(Document.embedding.cosine_distance(query_embedding))
            .limit(limit)
            .all()
        )

        # Sestavi response
        similar = []
        for doc, sim in results:
            # Če je dokument imel template, dodaj info
            template_info = None
            if doc.template_id and doc.template:
                template_info = {
                    "id": str(doc.template.id),
                    "name": doc.template.name,
                    "field_count": len(doc.template.template_fields),
                    "document_type": doc.template.document_type,
                }

            similar.append({
                "id": str(doc.id),
                "filename": doc.filename,
                "upload_date": doc.upload_date.isoformat(),
                "similarity": round(max(0.0, min(1.0, float(sim))), 3) if math.isfinite(float(sim or 0)) else 0.0,
                "template": template_info,
            })

        return {"similar_documents": similar}

    finally:
        os.remove(tmp_path)

@app.post("/api/batches")
async def create_batch(
    files: list[UploadFile] = File(...),
    name: str = Form(...),
    fields: str = Form(...),
    template_id: str = Form(""),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """
    Ustvari nov batch z eksplicitnimi polji.
    `fields` je obvezen JSON array — backend ne ugiba.
    `template_id` je opcijski (samo za sledenje).
    """
    if not files:
        return {"error": "Ni datotek za naložiti"}

    try:
        fields_list = json.loads(fields)
    except Exception:
        return {"error": "Neveljaven JSON za fields"}

    if not fields_list:
        return {"error": "Vsaj eno polje mora biti definirano"}

    # 1) Ustvari Batch zapis
    batch = Batch(
        user_id=user.id,
        name=name,
        template_id=template_id if template_id else None,
        status="processing",
        total_documents=len(files),
        completed_documents=0,
    )
    db.add(batch)
    db.flush()
    batch_id = str(batch.id)

    # 2) Shrani vse PDF-je in ustvari Document zapise
    document_ids = []
    for file in files:
        pdf_filename = f"{batch_id[:8]}_{file.filename}"
        pdf_path = os.path.join(PDF_DIR, pdf_filename)
        content = await file.read()
        with open(pdf_path, "wb") as f:
            f.write(content)

        text = extract_text(pdf_path)
        with fitz.open(pdf_path) as pdf_doc:
            total_pages = pdf_doc.page_count

        document = Document(
            user_id=user.id,
            filename=file.filename,
            pdf_path=pdf_path,
            total_pages=total_pages,
            document_text=text,
            batch_id=batch.id,
            template_id=batch.template_id,
            status="pending"
        )
        db.add(document)
        db.flush()
        document_ids.append(str(document.id))

    db.commit()
    print(f"USTVARJEN Batch: {batch_id} ({len(files)} datotek, {len(fields_list)} polj)")

    # 3) Zaženi background thread za procesiranje
    thread = threading.Thread(
        target=process_batch_background,
        args=(batch_id, document_ids, fields_list),
        daemon=True
    )
    thread.start()

    return {
        "batch_id": batch_id,
        "name": name,
        "total": len(files),
        "status": "processing"
    }


def process_batch_background(batch_id: str, document_ids: list, fields_list: list):
    """
    Background thread funkcija — procesira vse dokumente v batch-u z eksplicitnimi polji.
    """
    print(f"BATCH {batch_id[:8]}: začenjam procesiranje {len(document_ids)} dokumentov")
    db = SessionLocal()
    try:
        batch = db.query(Batch).filter(Batch.id == batch_id).first()
        if not batch:
            print(f"BATCH {batch_id[:8]}: ne obstaja")
            return

        # Procesira en dokument naenkrat
        for doc_id in document_ids:
            try:
                document = db.query(Document).filter(Document.id == doc_id).first()
                if not document:
                    continue

                # 1) Embedding
                document.embedding = generate_embedding(document.document_text)

                # 2) LLM ekstrakcija
                prompt = build_extract_prompt(fields_list, document.document_text)
                raw = call_llm(prompt)
                extraction_data = json.loads(raw)

                # 3) Najdi koordinate + confidence
                results = find_coordinates_and_confidence(document.pdf_path, extraction_data)

                confidences = [v["confidence"] for v in results.values() if v.get("confidence") is not None]
                avg_conf = sum(confidences) / len(confidences) if confidences else 0.0

                # 4) Shrani Extraction
                extraction = Extraction(
                    document_id=document.id,
                    model_used="openai/gpt-4o-mini",
                    avg_confidence=avg_conf,
                    raw_llm_response=extraction_data
                )
                db.add(extraction)
                db.flush()

                # 5) Shrani Fields
                for field_key, field_data in results.items():
                    field_record = Field(
                        extraction_id=extraction.id,
                        field_key=field_key,
                        field_value=field_data.get("value"),
                        source_text=field_data.get("source_text"),
                        confidence=field_data.get("confidence") or 0.0,
                        page_number=field_data.get("page"),
                        rectangles=field_data.get("rectangles", [])
                    )
                    db.add(field_record)

                # 6) Posodobi progress
                batch.completed_documents += 1
                db.commit()
                print(f"BATCH {batch_id[:8]}: {batch.completed_documents}/{batch.total_documents} — {document.filename}")

            except Exception as e:
                db.rollback()
                print(f"BATCH {batch_id[:8]}: NAPAKA pri {doc_id[:8]}: {e}")
                batch.completed_documents += 1
                db.commit()

        # Zaključi batch
        batch.status = "completed"
        db.commit()
        print(f"BATCH {batch_id[:8]}: KONČANO ({batch.completed_documents}/{batch.total_documents})")

    finally:
        db.close()


@app.get("/api/batches")
def list_batches(
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Seznam vseh batch-ov, sortiran po datumu."""
    batches = db.query(Batch).filter(
        Batch.user_id == user.id
    ).order_by(Batch.created_date.desc()).all()

    result = []
    for b in batches:
        template_name = None
        if b.template_id and b.template:
            template_name = b.template.name

        result.append({
            "id": str(b.id),
            "name": b.name,
            "created_date": b.created_date.isoformat(),
            "status": b.status,
            "total_documents": b.total_documents,
            "completed_documents": b.completed_documents,
            "template_id": str(b.template_id) if b.template_id else None,
            "template_name": template_name,
        })

    return {"batches": result}


@app.get("/api/batches/{batch_id}")
def get_batch(
    batch_id: str,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """
    Vrne batch z vsemi dokumenti in njihovimi rezultati.
    Frontend bo polled ta endpoint za progress in končne rezultate.
    """
    batch = db.query(Batch).filter(
        Batch.id == batch_id,
        Batch.user_id == user.id,
    ).first()
    if not batch:
        return {"error": "Batch ne obstaja"}

    template_name = None
    if batch.template_id and batch.template:
        template_name = batch.template.name

    documents = []
    for doc in batch.documents:
        latest_extraction = None
        if doc.extractions:
            latest_extraction = max(doc.extractions, key=lambda e: e.extraction_date)

        fields_summary = {}
        if latest_extraction:
            for field in latest_extraction.fields:
                fields_summary[field.field_key] = {
                    "value": field.field_value,
                    "confidence": field.confidence,
                }

        documents.append({
            "id": str(doc.id),
            "filename": doc.filename,
            "total_pages": doc.total_pages,
            "avg_confidence": latest_extraction.avg_confidence if latest_extraction else None,
            "fields": fields_summary,
            "has_extraction": latest_extraction is not None,
        })

    return {
        "id": str(batch.id),
        "name": batch.name,
        "created_date": batch.created_date.isoformat(),
        "status": batch.status,
        "total_documents": batch.total_documents,
        "completed_documents": batch.completed_documents,
        "template_id": str(batch.template_id) if batch.template_id else None,
        "template_name": template_name,
        "documents": documents,
    }


@app.delete("/api/batches/{batch_id}")
def delete_batch(
    batch_id: str,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Izbriše batch in vse povezane dokumente (cascade)."""
    batch = db.query(Batch).filter(
        Batch.id == batch_id,
        Batch.user_id == user.id,
    ).first()
    if not batch:
        return {"error": "Batch ne obstaja"}

    # Izbriši PDF datoteke z diska
    for doc in batch.documents:
        if os.path.exists(doc.pdf_path):
            try:
                os.remove(doc.pdf_path)
            except Exception:
                pass

    db.delete(batch)
    db.commit()

    return {"status": "izbrisano"}


@app.get("/api/batches/{batch_id}/export")
def export_batch(
    batch_id: str,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """
    Vrne raw podatke vseh dokumentov v batch-u.
    Samo field_key → value mapping, brez confidence ali metapodatkov.
    """
    batch = db.query(Batch).filter(
        Batch.id == batch_id,
        Batch.user_id == user.id,
    ).first()
    if not batch:
        return {"error": "Batch ne obstaja"}

    rows = []
    for doc in batch.documents:
        if not doc.extractions:
            continue
        latest = max(doc.extractions, key=lambda e: e.extraction_date)

        row = {"filename": doc.filename}
        for field in latest.fields:
            row[field.field_key] = field.field_value or ""

        rows.append(row)

    return {
        "batch_name": batch.name,
        "rows": rows,
    }



@app.get("/api/batches/{batch_id}/export-txt")
def export_batch_txt(
    batch_id: str,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """
    Vrne batch rezultate kot raw TXT — samo field: value pari.
    Brez confidence, korekcij ali metapodatkov.
    """
    batch = db.query(Batch).filter(
        Batch.id == batch_id,
        Batch.user_id == user.id,
    ).first()
    if not batch:
        return {"error": "Batch ne obstaja"}

    lines = []

    for doc in batch.documents:
        if not doc.extractions:
            continue
        latest = max(doc.extractions, key=lambda e: e.extraction_date)

        lines.append(f"=== {doc.filename} ===")
        for field in sorted(latest.fields, key=lambda f: f.field_key):
            value = field.field_value or ""
            lines.append(f"{field.field_key}: {value}")
        lines.append("")  # prazna vrstica med dokumenti

    content = "\n".join(lines)

    return StreamingResponse(
        io.BytesIO(content.encode("utf-8")),
        media_type="text/plain; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{_safe_filename("batch_" + batch.name, "txt")}"'}
    )


@app.get("/api/batches/{batch_id}/export-excel")
def export_batch_excel(
    batch_id: str,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Vrne batch rezultate kot Excel datoteko."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    batch = db.query(Batch).filter(
        Batch.id == batch_id,
        Batch.user_id == user.id,
    ).first()
    if not batch:
        return {"error": "Batch ne obstaja"}

    wb = Workbook()
    ws = wb.active
    ws.title = "Rezultati"

    # 1) Zberi vse unique field keys iz vseh dokumentov
    all_keys = set()
    for doc in batch.documents:
        if doc.extractions:
            latest = max(doc.extractions, key=lambda e: e.extraction_date)
            for field in latest.fields:
                all_keys.add(field.field_key)
    sorted_keys = sorted(all_keys)

    # 2) Header vrstica — samo Datoteka + polja, brez metapodatkov
    headers = ["Datoteka"] + sorted_keys
    ws.append(headers)

    # Stil header-ja
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="3B82F6")
        cell.alignment = Alignment(horizontal="left", vertical="center")

    # 3) Data vrstice
    for doc in batch.documents:
        if not doc.extractions:
            continue
        latest = max(doc.extractions, key=lambda e: e.extraction_date)
        field_map = {f.field_key: (f.field_value or "") for f in latest.fields}

        row = [doc.filename]
        for key in sorted_keys:
            row.append(field_map.get(key, ""))
        ws.append(row)

    # 4) Auto-size stolpci
    for col in ws.columns:
        max_len = max(len(str(c.value or '')) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    # 5) Stream kot Excel datoteka
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = _safe_filename(f"{batch.name}_{batch_id[:8]}", "xlsx")

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )